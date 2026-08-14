from __future__ import annotations

import base64
import hashlib
import hmac
import io
import os
import secrets
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path

import qrcode
from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from passlib.context import CryptContext
from starlette.middleware.sessions import SessionMiddleware

ROOT = Path(__file__).parent
DB_PATH = ROOT / "presensi.db"
SECRET = os.environ.get("APP_SECRET", "dev-secret-ganti-saat-produksi")
ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Mokaku2026!")
COMMITTEE_ACCESS_TOKEN = os.environ.get("COMMITTEE_ACCESS_TOKEN", "PanitiaMokaku2026!")
COMMITTEE_DIVISIONS = (
    "Penanggung Jawab", "Steering Committee", "Ketua Pelaksana",
    "Wakil Ketua Pelaksana", "Bendahara", "DPM", "Divisi Acara",
    "Divisi Humas", "Divisi Konsumsi", "Divisi Logistik", "Divisi Medis",
    "Divisi PDD", "Divisi Sponsor", "Divisi Teknisi Lapangan",
)
STUDY_PROGRAMS = (
    "Biologi", "Kimia", "Fisika", "Matematika", "Ilmu Komputer", "IPSE",
    "Pendidikan Biologi", "Pendidikan Kimia", "Pendidikan Fisika",
    "Pendidikan Matematika", "Pendidikan Ilmu Komputer",
)
pwd = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
app = FastAPI(title="Presensi MOKAKU FPMIPA 2026")
app.add_middleware(SessionMiddleware, secret_key=SECRET, same_site="lax", https_only=False)
app.mount("/static", StaticFiles(directory=ROOT / "static"), name="static")

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False

USE_PG = bool(DATABASE_URL and HAS_PSYCOPG2)

if USE_PG:
    DBIntegrityError = (psycopg2.IntegrityError, sqlite3.IntegrityError)
else:
    DBIntegrityError = (sqlite3.IntegrityError,)

class PGConnWrapper:
    def __init__(self, conn):
        self.conn = conn
        self.lastrowid = None

    def execute(self, sql, params=()):
        sql_pg = sql.replace("?", "%s")
        cur = self.conn.cursor(cursor_factory=RealDictCursor)
        if "INSERT INTO users(" in sql and "RETURNING id" not in sql_pg:
            sql_pg += " RETURNING id"
        cur.execute(sql_pg, params)
        if "RETURNING id" in sql_pg:
            row = cur.fetchone()
            if row:
                self.lastrowid = row["id"]
        return cur

    def executescript(self, sql):
        sql_pg = sql.replace("?", "%s")
        cur = self.conn.cursor()
        cur.execute(sql_pg)
        return cur

    def commit(self):
        self.conn.commit()

    def close(self):
        self.conn.close()

@contextmanager
def db():
    if USE_PG:
        conn = psycopg2.connect(DATABASE_URL)
        wrapper = PGConnWrapper(conn)
        try:
            yield wrapper
            wrapper.commit()
        finally:
            wrapper.close()
    else:
        con = sqlite3.connect(DB_PATH)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA foreign_keys=ON")
        try:
            yield con
            con.commit()
        finally:
            con.close()

def init_db():
    if USE_PG:
        with db() as con:
            con.execute("""
            CREATE TABLE IF NOT EXISTS users(
              id SERIAL PRIMARY KEY,
              nim TEXT UNIQUE NOT NULL,
              name TEXT NOT NULL,
              email TEXT UNIQUE NOT NULL,
              password_hash TEXT NOT NULL,
              qr_nonce TEXT UNIQUE NOT NULL,
              created_at TEXT NOT NULL,
              role TEXT NOT NULL DEFAULT 'peserta',
              committee_division TEXT,
              study_program TEXT
            );
            CREATE TABLE IF NOT EXISTS sessions(
              id SERIAL PRIMARY KEY,
              name TEXT NOT NULL,
              week INTEGER NOT NULL,
              session_date TEXT NOT NULL,
              is_open INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL,
              audience TEXT NOT NULL DEFAULT 'peserta'
            );
            CREATE TABLE IF NOT EXISTS attendance(
              id SERIAL PRIMARY KEY,
              user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
              session_id INTEGER NOT NULL REFERENCES sessions(id) ON DELETE CASCADE,
              scanned_at TEXT NOT NULL,
              scanned_by TEXT NOT NULL,
              source TEXT NOT NULL DEFAULT 'qr',
              UNIQUE(user_id, session_id)
            );
            """)
            for audience, name in (("peserta", "Kehadiran Peserta"), ("panitia", "Kehadiran Panitia")):
                res = con.execute("SELECT id FROM sessions WHERE audience=? LIMIT 1", (audience,)).fetchone()
                if not res:
                    con.execute("INSERT INTO sessions(name,week,session_date,is_open,created_at,audience) VALUES(?,?,?,?,?,?)",
                                (name, 1, date.today().isoformat(), 1, datetime.now().isoformat(timespec="seconds"), audience))
    else:
        with db() as con:
            con.executescript("""
            CREATE TABLE IF NOT EXISTS users(
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              nim TEXT UNIQUE NOT NULL,
              name TEXT NOT NULL,
              email TEXT UNIQUE NOT NULL,
              password_hash TEXT NOT NULL,
              qr_nonce TEXT UNIQUE NOT NULL,
              created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sessions(
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL,
              week INTEGER NOT NULL,
              session_date TEXT NOT NULL,
              is_open INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attendance(
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
              session_id INTEGER NOT NULL REFERENCES sessions(id) ON DELETE CASCADE,
              scanned_at TEXT NOT NULL,
              scanned_by TEXT NOT NULL,
              source TEXT NOT NULL DEFAULT 'qr',
              UNIQUE(user_id, session_id)
            );
            """)
            user_columns = {row[1] for row in con.execute("PRAGMA table_info(users)").fetchall()}
            if "role" not in user_columns:
                con.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'peserta'")
            if "committee_division" not in user_columns:
                con.execute("ALTER TABLE users ADD COLUMN committee_division TEXT")
            if "study_program" not in user_columns:
                con.execute("ALTER TABLE users ADD COLUMN study_program TEXT")
            session_columns = {row[1] for row in con.execute("PRAGMA table_info(sessions)").fetchall()}
            if "audience" not in session_columns:
                con.execute("ALTER TABLE sessions ADD COLUMN audience TEXT NOT NULL DEFAULT 'peserta'")
            columns = {row[1] for row in con.execute("PRAGMA table_info(attendance)").fetchall()}
            if "source" not in columns:
                con.execute("ALTER TABLE attendance ADD COLUMN source TEXT NOT NULL DEFAULT 'qr'")
            for audience, name in (("peserta", "Kehadiran Peserta"), ("panitia", "Kehadiran Panitia")):
                existing = con.execute("SELECT id FROM sessions WHERE audience=? LIMIT 1", (audience,)).fetchone()
                if not existing:
                    con.execute("INSERT INTO sessions(name,week,session_date,is_open,created_at,audience) VALUES(?,?,?,?,?,?)",
                                (name, 1, date.today().isoformat(), 1, datetime.now().isoformat(timespec="seconds"), audience))

@app.on_event("startup")
def startup(): init_db()

def flash(request: Request, message: str, kind: str = "ok"):
    request.session["flash"] = {"message": message, "kind": kind}

def participant(request: Request):
    uid = request.session.get("participant_id")
    if not uid: return None
    with db() as con: return con.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()

def is_admin(request: Request): return bool(request.session.get("admin"))

def sign_qr(user):
    payload = f"MOKAKU26:{user['nim']}:{user['qr_nonce']}"
    signature = hmac.new(SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:24]
    return f"{payload}:{signature}"

def parse_qr(token: str):
    parts = token.strip().split(":")
    if len(parts) != 4 or parts[0] != "MOKAKU26": return None
    payload = ":".join(parts[:3])
    expected = hmac.new(SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:24]
    if not hmac.compare_digest(parts[3], expected): return None
    return {"nim": parts[1], "nonce": parts[2]}

def qr_data_url(value: str):
    image = qrcode.make(value)
    out = io.BytesIO(); image.save(out, format="PNG")
    return "data:image/png;base64," + base64.b64encode(out.getvalue()).decode()

def render(request: Request, template: str, **context):
    from fastapi.templating import Jinja2Templates
    templates = Jinja2Templates(directory=ROOT / "templates")
    context.update(request=request, flash=request.session.pop("flash", None), participant=participant(request), admin=is_admin(request))
    return templates.TemplateResponse(request, template, context)

@app.get("/", response_class=HTMLResponse)
def home(request: Request): return render(request, "home.html")

@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request): return render(request, "register.html", account_role="peserta", study_programs=STUDY_PROGRAMS)

@app.get("/register-panitia", response_class=HTMLResponse)
def register_committee_page(request: Request): return render(request, "register.html", account_role="panitia", committee_divisions=COMMITTEE_DIVISIONS)

@app.post("/register")
def register(request: Request, nim: str = Form(...), name: str = Form(...), email: str = Form(...), role: str = Form("peserta"), committee_division: str = Form(""), study_program: str = Form(""), access_token: str = Form("")):
    nim, name, email = nim.strip(), name.strip(), email.strip().lower()
    role = role if role in ("peserta", "panitia") else "peserta"
    committee_division = committee_division.strip()
    study_program = study_program.strip()
    access_token = access_token.strip()

    if role == "panitia":
        if not secrets.compare_digest(access_token, COMMITTEE_ACCESS_TOKEN):
            flash(request, "Token akses panitia tidak valid.", "error")
            return RedirectResponse("/register-panitia", 303)
        if committee_division not in COMMITTEE_DIVISIONS:
            flash(request, "Pilih jabatan atau divisi panitia yang tersedia.", "error")
            return RedirectResponse("/register-panitia", 303)
    if role == "peserta" and study_program not in STUDY_PROGRAMS:
        flash(request, "Pilih program studi yang tersedia.", "error")
        return RedirectResponse("/register", 303)
    if role == "peserta": committee_division = ""
    if role == "panitia": study_program = ""
    if not nim.isdigit() or len(nim) < 5:
        flash(request, "NIM harus berupa angka minimal 5 digit.", "error")
        return RedirectResponse("/register", 303)
    try:
        with db() as con:
            cur = con.execute("INSERT INTO users(nim,name,email,password_hash,qr_nonce,created_at,role,committee_division,study_program) VALUES(?,?,?,?,?,?,?,?,?)",
                (nim, name, email, "EMAIL_NIM_LOGIN", secrets.token_urlsafe(12), datetime.now().isoformat(timespec="seconds"), role, committee_division or None, study_program or None))
            request.session["participant_id"] = cur.lastrowid
        flash(request, "Akun berhasil dibuat. QR presensi kamu sudah siap.")
        return RedirectResponse("/participant", 303)
    except DBIntegrityError:
        flash(request, "NIM atau email sudah terdaftar.", "error")
        return RedirectResponse("/register", 303)

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request): return render(request, "login.html", login_role="peserta")

@app.get("/login-panitia", response_class=HTMLResponse)
def login_committee_page(request: Request): return render(request, "login.html", login_role="panitia")

@app.post("/login")
@app.post("/login-panitia")
def login(request: Request, nim: str = Form(...), email: str = Form(...)):
    expected_role = "panitia" if request.url.path == "/login-panitia" else "peserta"
    with db() as con: user = con.execute("SELECT * FROM users WHERE nim=? AND email=? AND role=?", (nim.strip(), email.strip().lower(), expected_role)).fetchone()
    if not user:
        flash(request, "Email atau NIM tidak cocok.", "error"); return RedirectResponse(request.url.path, 303)
    request.session.clear(); request.session["participant_id"] = user["id"]
    return RedirectResponse("/participant", 303)

@app.get("/participant", response_class=HTMLResponse)
def participant_dashboard(request: Request):
    user = participant(request)
    if not user: return RedirectResponse("/login", 303)
    with db() as con:
        history = con.execute("""SELECT s.name,s.week,s.session_date,a.scanned_at FROM attendance a
          JOIN sessions s ON s.id=a.session_id WHERE a.user_id=? ORDER BY a.scanned_at DESC""", (user["id"],)).fetchall()
    return render(request, "participant.html", user=user, qr=qr_data_url(sign_qr(user)), qr_token=sign_qr(user), history=history)

@app.get("/logout")
def logout(request: Request): request.session.clear(); return RedirectResponse("/", 303)

@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_page(request: Request): return render(request, "admin_login.html")

@app.post("/admin/login")
def admin_login(request: Request, username: str = Form(...), password: str = Form(...)):
    if secrets.compare_digest(username, ADMIN_USER) and secrets.compare_digest(password, ADMIN_PASSWORD):
        request.session.clear(); request.session["admin"] = True; return RedirectResponse("/admin", 303)
    flash(request, "Kredensial admin salah.", "error"); return RedirectResponse("/admin/login", 303)

@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request):
    if not is_admin(request): return RedirectResponse("/admin/login", 303)
    with db() as con:
        sessions = con.execute("""SELECT s.*,COUNT(a.id) attendance_count FROM sessions s
          LEFT JOIN attendance a ON a.session_id=s.id GROUP BY s.id ORDER BY s.session_date DESC,s.week DESC""").fetchall()
        users = con.execute("SELECT COUNT(*) AS total FROM users WHERE role='peserta'").fetchone()["total"]
        committee = con.execute("SELECT COUNT(*) AS total FROM users WHERE role='panitia'").fetchone()["total"]
        total = con.execute("SELECT COUNT(*) AS total FROM attendance").fetchone()["total"]
        recent = con.execute("""SELECT u.nim,u.name,s.name session_name,a.scanned_at FROM attendance a
          JOIN users u ON u.id=a.user_id JOIN sessions s ON s.id=a.session_id ORDER BY a.scanned_at DESC LIMIT 8""").fetchall()
    return render(request, "admin.html", sessions=sessions, user_count=users, committee_count=committee, attendance_count=total, recent=recent)

@app.post("/admin/session")
def create_session(request: Request, name: str = Form(...), week: int = Form(...), session_date: str = Form(...), audience: str = Form("peserta")):
    if not is_admin(request): return RedirectResponse("/admin/login", 303)
    audience = audience if audience in ("peserta", "panitia") else "peserta"
    with db() as con: con.execute("INSERT INTO sessions(name,week,session_date,is_open,created_at,audience) VALUES(?,?,?,?,?,?)", (name.strip(),week,session_date,1,datetime.now().isoformat(timespec="seconds"),audience))
    flash(request, "Sesi presensi berhasil dibuat."); return RedirectResponse("/admin",303)

@app.post("/admin/session/{session_id}/toggle")
def toggle_session(session_id: int, request: Request):
    if not is_admin(request): return RedirectResponse("/admin/login",303)
    with db() as con: con.execute("UPDATE sessions SET is_open=CASE is_open WHEN 1 THEN 0 ELSE 1 END WHERE id=?",(session_id,))
    return RedirectResponse("/admin",303)

@app.get("/admin/scan/{session_id}", response_class=HTMLResponse)
def scan_page(session_id: int, request: Request):
    if not is_admin(request): return RedirectResponse("/admin/login",303)
    with db() as con: session=con.execute("SELECT * FROM sessions WHERE id=?",(session_id,)).fetchone()
    if not session: return RedirectResponse("/admin",303)
    return render(request,"scan.html",session=session)

@app.post("/admin/api/scan/{session_id}")
async def record_scan(session_id: int, request: Request):
    if not is_admin(request): return {"ok":False,"message":"Sesi admin berakhir."}
    body=await request.json(); parsed=parse_qr(str(body.get("token", "")))
    if not parsed: return {"ok":False,"message":"QR tidak valid atau bukan QR MOKAKU."}
    with db() as con:
        session=con.execute("SELECT * FROM sessions WHERE id=?",(session_id,)).fetchone()
        user=con.execute("SELECT * FROM users WHERE nim=? AND qr_nonce=?",(parsed["nim"],parsed["nonce"])).fetchone()
        if not session or not session["is_open"]: return {"ok":False,"message":"Sesi presensi sudah ditutup."}
        if not user: return {"ok":False,"message":"Peserta tidak ditemukan."}
        if user["role"] != session["audience"]: return {"ok":False,"message":f"QR ini milik {user['role']}, bukan untuk sesi {session['audience']}."}
        try:
            now=datetime.now().isoformat(timespec="seconds")
            con.execute("INSERT INTO attendance(user_id,session_id,scanned_at,scanned_by,source) VALUES(?,?,?,?,?)",(user["id"],session_id,now,ADMIN_USER,"QR Scanner"))
            identity_detail = user["study_program"] if user["role"] == "peserta" else user["committee_division"]
            return {"ok":True,"message":f"Hadir tercatat: {user['name']} ({user['nim']})","name":user["name"],"nim":user["nim"],"role":user["role"],"study_program":user["study_program"],"committee_division":user["committee_division"],"identity_detail":identity_detail or "-","time":now[11:19]}
        except DBIntegrityError: return {"ok":False,"duplicate":True,"message":f"{user['name']} sudah presensi di sesi ini."}

@app.post("/admin/manual/{session_id}")
def manual_attendance(session_id: int, request: Request, nim: str=Form(...)):
    if not is_admin(request): return RedirectResponse("/admin/login",303)
    with db() as con:
        session=con.execute("SELECT * FROM sessions WHERE id=?",(session_id,)).fetchone(); user=con.execute("SELECT * FROM users WHERE nim=?",(nim.strip(),)).fetchone()
        if not session or not session["is_open"]: flash(request,"Sesi tidak tersedia atau sudah ditutup.","error")
        elif not user: flash(request,"NIM peserta tidak ditemukan.","error")
        elif user["role"] != session["audience"]: flash(request,f"Akun ini berjenis {user['role']}, bukan untuk sesi {session['audience']}.","error")
        else:
            try: con.execute("INSERT INTO attendance(user_id,session_id,scanned_at,scanned_by,source) VALUES(?,?,?,?,?)",(user["id"],session_id,datetime.now().isoformat(timespec="seconds"),ADMIN_USER,"Input Manual")); flash(request,f"Presensi {user['name']} berhasil dicatat.")
            except DBIntegrityError: flash(request,"Peserta sudah presensi di sesi ini.","error")
    return RedirectResponse(f"/admin/scan/{session_id}",303)

def build_excel_export(audience_filter: str | None = None):
    audience_filter = audience_filter if audience_filter in ("peserta", "panitia") else None
    with db() as con:
        if audience_filter:
            users=con.execute("SELECT * FROM users WHERE role=? ORDER BY nim",(audience_filter,)).fetchall()
            sessions=con.execute("SELECT * FROM sessions WHERE audience=? ORDER BY week,session_date",(audience_filter,)).fetchall()
            attendance_rows=con.execute("""SELECT a.*,u.nim,u.name participant_name,u.email,u.role,u.committee_division,u.study_program,s.name session_name,s.week,s.session_date
              FROM attendance a JOIN users u ON u.id=a.user_id JOIN sessions s ON s.id=a.session_id
              WHERE u.role=? AND s.audience=? ORDER BY a.scanned_at DESC""",(audience_filter,audience_filter)).fetchall()
        else:
            users=con.execute("SELECT * FROM users ORDER BY role,nim").fetchall()
            sessions=con.execute("SELECT * FROM sessions ORDER BY audience,week,session_date").fetchall()
            attendance_rows=con.execute("""SELECT a.*,u.nim,u.name participant_name,u.email,u.role,u.committee_division,u.study_program,s.name session_name,s.week,s.session_date
          FROM attendance a JOIN users u ON u.id=a.user_id JOIN sessions s ON s.id=a.session_id
          ORDER BY a.scanned_at DESC""").fetchall()
        records={(r["user_id"],r["session_id"]):r for r in attendance_rows}
    wb=Workbook(); ws=wb.active; ws.title="Rekap Presensi"
    headers=["NIM","Nama","Email","Jenis Akun","Program Studi","Jabatan/Divisi Panitia"]+[s['name'] for s in sessions]+["Total Hadir","Persentase","Nilai Presensi"]
    ws.append(headers)
    for u in users:
        marks=["Hadir" if u["role"]==s["audience"] and (u["id"],s["id"]) in records else ("" if u["role"]==s["audience"] else "N/A") for s in sessions]; relevant=[m for m in marks if m!="N/A"]; total=relevant.count("Hadir"); pct=round(total/len(relevant)*100,2) if relevant else 0
        ws.append([u["nim"],u["name"],u["email"],u["role"].title(),u["study_program"] or "-",u["committee_division"] or "-",*marks,total,pct,pct])
    ws.freeze_panes="G2"; ws.auto_filter.ref=ws.dimensions
    log=wb.create_sheet("Log Kehadiran Real-Time")
    log.append(["No","Timestamp Scan","Tanggal","Waktu","NIM","Nama","Email","Jenis Akun","Program Studi","Jabatan/Divisi Panitia","Pekan","Sesi","Tanggal Sesi","Metode","Dicatat Oleh"])
    for number,row in enumerate(attendance_rows,1):
        timestamp=datetime.fromisoformat(row["scanned_at"])
        log.append([number,timestamp,timestamp.date(),timestamp.time(),row["nim"],row["participant_name"],row["email"],row["role"].title(),row["study_program"] or "-",row["committee_division"] or "-",row["week"],row["session_name"],row["session_date"],row["source"],row["scanned_by"]])
    log.freeze_panes="A2"; log.auto_filter.ref=log.dimensions
    for sheet in (ws,log):
        for cell in sheet[1]:
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor="17365D")
            cell.alignment=Alignment(horizontal="center")
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or "")) for c in col)+2,40)
    for cell in log["B"][1:]: cell.number_format="dd/mm/yyyy hh:mm:ss"
    for cell in log["C"][1:]: cell.number_format="dd/mm/yyyy"
    for cell in log["D"][1:]: cell.number_format="hh:mm:ss"
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out

def excel_response(out: io.BytesIO, audience: str):
    filename=f"rekap-absensi-{audience}-mokaku-fpmipa-2026-{date.today().isoformat()}.xlsx"
    return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{filename}"'})

@app.get("/admin/export-peserta.xlsx")
def export_participants_excel(request: Request):
    if not is_admin(request): return RedirectResponse("/admin/login",303)
    return excel_response(build_excel_export("peserta"),"peserta")

@app.get("/admin/export-panitia.xlsx")
def export_committee_excel(request: Request):
    if not is_admin(request): return RedirectResponse("/admin/login",303)
    return excel_response(build_excel_export("panitia"),"panitia")

@app.get("/admin/export.xlsx")
def export_excel(request: Request):
    if not is_admin(request): return RedirectResponse("/admin/login",303)
    return excel_response(build_excel_export(),"gabungan")

init_db()
